import pandas as pd
import datetime
import openpyxl
from pandas import DataFrame

from globals import *  # - импорт всех переменных (предпочтительнее)
from pandas.io.excel import ExcelWriter


# ============================================== Сохраняем протокол ====================================================
def save_protocol(compound_name_files: str, corp_name_prefix: str, path_save_protocol: str):
    """
    Функция записи "идентификатора загрузки" дампа (уникальное имя файла и путь) в протокол при сохранении дампа.
    Данный механизм нужен для того, что бы идентифицировать дампы от разных таблиц при загрузке обратно в оперативную
    память.

    :param compound_name_fails: Сохраняемое в протоколе комбинированное имя файла ("идентификатора загрузки"),
    defaults to None
    :type compound_name_fails: str

    :param corp_name_prefix: Имя корпорации, будет использовано, как имя листа, defaults to None
    :type corp_name_prefix: str

    :param path_save_protocol: Относительный путь и имя файла протокола, defaults to = r'\data\save_damp'
    :type path_save_protocol: str

    :rtype: Object
    :return: compound_name_files

    :notes:
    # Делаем compound_name_fails обязательным параметром, т.к. переменная хранит идентификатор загрузки.
    Логика: Функция сохраняет запись в файл эксель об адресе и месте сохранения (в виде составного имени файла, \
    \ включающего в себя путь и имя - он же "идентификатора загрузки").
    Процессы: 1й - Сохраняем составное (на основе входных данных) имя в лист,
    аналогично сохраняем в другой лист дату и время. Далее оба листа помещаем в дата фрейм, как 2 колонки.
    Предполагается сохранение разных идентификаторов на разные листы (будут создаваться)

    """

    # 0. Подготовка данных:
    # -----------------------------------------Сохраняем в переменной полный путь --------------------------------------
    path_save_protocol_file = (path_save_protocol + 'protocol.xlsx')  # - Собираем полный путь до файла.
    #  path_save_protocol_file = содержит полный путь (по умолчанию папка сохранения дампа).
    # 1. Проверка существования файла: Если нет (False), то пишем
    not_none_protocol = os.path.isfile(path_save_protocol_file)
    # if not_none_protocol is not True:  # Если файла нет.
    # ------------------------------------------------------------------------------------------------------------------
    # 2. Сохраняем в файл эксель протокол.
    # -------------------------------------- Записываем данные о сохранении в протокол ---------------------------------
    files_name_list: list = [compound_name_files]  # Помещаем в лист нашу будущую запись-идентификатор \
    # для протокола (сохранения/загрузки).
    data_save_protocol: object = datetime.datetime.now()  # - Присваиваем дату + время сейчас \
    # и записываем ее в переменную,
    data_save_list: list = [data_save_protocol]  # Помещаем в лист переменную с записью даты.
    # ---------------------------------------- Сохраняем данные в дата фрем --------------------------------------------
    df_protocol_to_excel: DataFrame = pd.DataFrame({'full path to fails': files_name_list, 'data_save': data_save_list})
    # ---------------------------------------- Сохраняем файл эксель, он же протокол -----------------------------------

    if os.path.isfile(path_save_protocol_file) == True:  # Если файл существует,тогда:
        book = openpyxl.load_workbook(path_save_protocol_file)
        if corp_name_prefix in book.sheetnames:  # Если существует страница:
            # if pd.read_excel(dir_save_protocol, engine="openpyxl", sheet_name=i) == True: # с этим методом надо\
            # использовать try/except:
            # тогда читаем ее и записываем объединенный датафрейм (делаем перезапись листа)
            df_old = pd.read_excel(path_save_protocol_file, engine="openpyxl", sheet_name=corp_name_prefix, index_col=0)
            concat = pd.concat([df_old, df_protocol_to_excel], axis=0,
                               ignore_index=True)  # соединение датафреймов вертикально
            with (ExcelWriter(path_save_protocol_file, engine="openpyxl", mode="a",
                              if_sheet_exists="replace") as writer):
                concat.to_excel(writer, sheet_name=corp_name_prefix, index=True, index_label='id')
        else:  # Если не существует страницы:
            with (ExcelWriter(path_save_protocol_file, engine="openpyxl", mode="a") as writer):
                df_protocol_to_excel.to_excel(writer, sheet_name=corp_name_prefix, index=True, index_label='id')
    else:  # Если файла не существует,тогда:
        with (ExcelWriter(path_save_protocol_file, engine="openpyxl", mode="w") as writer):
            df_protocol_to_excel.to_excel(writer, sheet_name=corp_name_prefix, index=True, index_label='id')

    return print(f'Идентификатор загрузки дампа {corp_name_prefix} внесен в протокол')


# =============================================== Читаем протокол ======================================================

def read_protocol(path_save_protocol: str = None, corp_name_prefix: str = '') -> object:
    """
    Функция чтения "идентификатора загрузки" дампа (уникальное имя файла и путь) в протоколе.

    :param dir_seve_protocol: Сохраняемое в протоколе комбинированное имя файла ("идентификатора загрузки"),
    defaults to None
    :type dir_seve_protocol: str

    :param corp_name_prefix: Имя корпорации, будет использовано, как имя листа, defaults to None
    :type corp_name_prefix: str

    :rtype: Object
    :return: compound_name_fails
    :notes:
    Логика: на вход имя листа (обязательно, без этого не прочтет)
    """
    path_save_protocol_file = (path_save_protocol + 'protocol.xlsx')  # - Собираем полный путь до файла.
    if os.path.isfile(path_save_protocol_file) == True:  # Если файл существует,тогда читаем:
        book = openpyxl.load_workbook(path_save_protocol_file)
        if corp_name_prefix in book.sheetnames:  # Если существует страница с таким именем:
            load_df: DataFrame = pd.read_excel(path_save_protocol_file, engine="openpyxl", sheet_name=corp_name_prefix)
            # ---------------------------------------------------------------------------------------------------------
            range_id = load_df['id'].idxmax()  # максимальное значение для id работает (оследняя загрузка)
            load_path_damp = (load_df['full path to fails'].loc[range_id])  # работает - эквивалент полной функции
            #  ---------------------------------------------------------------------------------------------------------
            if os.path.isfile(load_path_damp) == False:  # Если файл не существует,тогда:
                i = range_id
                while i >= 0:
                    range_id = i - 1
                    if os.path.isfile(load_path_damp) == False:
                        continue
                    else:
                        break
                    dumps = load(load_path_damp)  # Тогда загружаем работает
                print('Загружен предыдущий дамп в протоколе.')
                return dumps
            else:  # Если удалось найти файл, тогда читаем запись
                # continue
                dumps = load(load_path_damp)  # Тогда загружаем работает
                print(f'Дамп {corp_name_prefix} успешно загружен!')
                return dumps
        else:
            print(f'Невозможно загрузить дамп: отсутствует лист с именем {corp_name_prefix} в протоколе!')
    else:
        print(f'Невозможно загрузить дамп: отсутствует файл протокола!')
    return dumps

# def decor_read_protocol(func):
#     def wrapper(path_save_protocol: str, corp_name_prefix: str=c):
#         for c in input_list_url:  # Переборка циклом значений словаря
#         load_path_damp:object = func(path_save_protocol, corp_name_prefix)
#         return load_path_damp
#     return wrapper
